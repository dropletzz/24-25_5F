# File excel
FILE_NAME = "_risposte.xlsx"
SHEET_NAME = "Risposte del modulo 1"
OUTPUT_SHEET_NAME = "risultati"

# Database
DB_HOST="0.0.0.0"
DB_USER="root"
DB_NAME="bookstore"
DB_PASS=""

# Risposte corrette
ANSWERS = [
#0 Selezionare nome e cognome di tutti gli utenti.
#(name, surname)
"SELECT name, surname FROM users;",

#1 Selezionare id, nome e cognome di tutti gli utenti che si chiamano "Mario"
#(name, surname)
"SELECT id, name, surname FROM users WHERE name = 'Mario' OR name = 'Marco';",

#2 Selezionare identificativo ed email di tutti gli utenti la cui mail termina con "@libero.it".
#(id, email)
"""
SELECT id, email FROM users
WHERE email LIKE "%@libero.it";
""",

#3 Selezionare i libri con costo inferiore a 15 euro, insieme al relativo autore.
#(title, name, surname, price)
"""
SELECT title, name, surname, price
FROM books JOIN authors ON author_id = authors.id
WHERE price < 15;
""",

#4 Selezionare tutti i generi insieme al conteggio dei libri per ogni genere.
#(genre, book_count)
"""
SELECT genre, COUNT(*) AS book_count
FROM books
GROUP BY genre;
""",

#5 Selezionare tutti gli autori insieme al conteggio dei libri che hanno scritto, compresi gli autori che non hanno scritto libri.
#(id, name, surname, book_count)
"""
SELECT a.id, a.name, a.surname, COUNT(b.id) AS book_count
FROM authors a LEFT JOIN books b ON b.author_id = a.id
GROUP BY a.id;
""",

#6 Selezionare tutti i libri insieme al loro rating medio (il rating sta nella tabella "reviews"). Ordinare i risultati partendo dal libro con il rating piu' alto.
#(id, title, average_rating)
"""
SELECT b.id, b.title, AVG(r.rating) AS average_rating
FROM books b LEFT JOIN reviews r ON b.id = r.book_id
GROUP BY b.id
ORDER BY average_rating DESC;
""",

#7 Abbiamo notato che tra gli utenti dell'applicazione ci sono dei bot che hanno scritto delle recensioni. I bot hanno la mail registrata su "@sharklasers.com". Scrivere una query che cancella dal database tutte le recensioni scritte da questi utenti.
# """
# DELETE FROM reviews
# WHERE user_id IN (
#   SELECT id FROM users
#   WHERE email LIKE "%@sharklasers.com"
# );
# """,
None, # non corretta automaticamente

#8 Selezionare tutti gli utenti che hanno speso almeno 50 euro negli ultimi 7 giorni.
#(id, email, total)
"""
SELECT u.id, u.email, SUM(bo.quantity * bo.price) AS total
FROM users u JOIN orders o ON u.id = o.user_id
JOIN books_orders bo ON o.id = bo.order_id
WHERE DATEDIFF(NOW(), o.created_at) <= 7
GROUP BY u.id
HAVING total >= 50;
""",

#9 Selezionare il costo medio degli ordini per ogni utente. Includere anche gli utenti che non hanno mai speso niente, i quali devono avere la media impostata a 0 (si veda la funzione IFNULL).
#(id, email, average_order_cost)
"""
SELECT u.id, u.email, AVG(IFNULL(order_cost, 0)) AS average_order_cost
FROM (
  -- calcola il costo di tutti gli ordini
  SELECT o.user_id, SUM(quantity * price) AS order_cost
  FROM orders o JOIN books_orders bo ON o.id = bo.order_id
  GROUP BY o.id
) AS oc RIGHT JOIN users u ON oc.user_id = u.id
GROUP BY u.id;
""",

#10 Vogliamo offrire un buono sconto del 50% sul libro "La gioia di scrivere" di Wislawa Szymborska, ma solo agli utenti che hanno acquistato libri di poesia almeno due volte (in due diversi ordini). Scrivere le query che generano i buoni sconto.
# """
# -- seleziona gli utenti a cui mandare il buono sconto
# SELECT u.id, COUNT(o.id) AS orders_count
# FROM users u JOIN orders o ON u.id = o.user_id
# WHERE o.id IN (
#   -- id degli ordini in cui c'e' almeno un libro di poesia
# 	SELECT o.id FROM books b
# 	JOIN books_orders bo ON b.id = bo.book_id
# 	JOIN orders o ON bo.order_id = o.id
# 	WHERE b.genre = "poesia"
# 	GROUP BY o.id
# )
# GROUP BY u.id
# HAVING orders_count >= 2;
# """
None, # non corretta automaticamente
]

DEBUG = False

################################
# FINE PARTE DI CONFIGURAZIONE #
################################


def log(x):
    if DEBUG:
        print(x)

import mariadb
from openpyxl import load_workbook

# connessione al db
connection = mariadb.connect(
    host=DB_HOST,
    user=DB_USER,
    database=DB_NAME,
    password=DB_PASS
)
cursor = connection.cursor()
print("Connessione al db riuscita")

# Confronta la risposta con quella attesa, restituisce:
# * 1 = risposta giusta
# * lettera = risposta errata (o parzialmente errata)
def compare_answers(correct, attempt):
    if correct is None:
        return "C"
    if attempt is None:
        return "A"

    def try_query(query, errcode):
        try:
            cursor.execute(query)
        except:
            log(f"ERROR RUNNING QUERY:\n{query}")
            return (None, errcode)
        try:
            res = cursor.fetchall()
        except:
            log(f"ERROR FETCHING RESULTS:\n{query}")
            return (None, f"{errcode}F")
        return (res, None)

    (cres, err) = try_query(correct, "C")
    if err:
        return err

    (res, err) = try_query(attempt, "A")
    if err:
        return err

    if len(cres) != len(res):
        return "L"

    for (cr, r) in zip(cres, res):
        if (cr != r):
            return "E"

    return 1

# dalla mail tira fuori cognome e iniziale del nome
# formati mail: "mrossi@qualcosa.it" o "rossi.m@qualcosa.it
def mail_to_name(mail):
    name = mail.split('@')[0]
    name_split = name.split(".")
    if (len(name_split) > 1):
        lastn = name_split[0]
        firstn = name_split[1]
    else:
        lastn = name[1:]
        firstn = name[0]
    return [ firstn, lastn ]


# Load the workbook
wb = load_workbook(FILE_NAME)

# Get a specific sheet by name
sheet = wb[SHEET_NAME]
if OUTPUT_SHEET_NAME in wb:
    out = wb[OUTPUT_SHEET_NAME]
else:
    out = wb.create_sheet(OUTPUT_SHEET_NAME)

i = 0
cell_value = sheet.cell(row=2, column=2).value
while cell_value is not None:
    i += 1
    cell_value = sheet.cell(row=2+i, column=2).value
student_count = i

for j in range(len(ANSWERS)):
    out.cell(row=1, column=3+j).value = f"#{j}"


for i in range(student_count):
    student_mail = sheet.cell(row=2+i, column=2).value
    [ firstn, lastn ] = mail_to_name(student_mail)
    out.cell(row=2+i, column=1).value = lastn
    out.cell(row=2+i, column=2).value = firstn
    print(f"Correggendo: {lastn} {firstn}")

    for j in range(len(ANSWERS)):
        student_answer = sheet.cell(row=2+i, column=3+j).value
        res = compare_answers(ANSWERS[j], student_answer)
        out.cell(row=2+i, column=3+j).value = res

wb.save(FILE_NAME)

print("Corretto tutto e aggiornato il file excel ;)")


