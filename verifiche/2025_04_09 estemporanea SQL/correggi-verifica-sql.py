# File excel
FILE_NAME = "_risposte.xlsx"
SHEET_NAME = "Risposte del modulo 1"
OUTPUT_SHEET_NAME = "risultati"

# Database
DB_HOST="0.0.0.0"
DB_USER="root"
DB_NAME="traffico"
DB_PASS=""

# Risposte corrette
ANSWERS = [
# a)
"""
SELECT s.nome AS nome_stazione, s.latitudine, s.longitudine, a.toponimo AS toponimo_strada
FROM StazioneDiMisurazione s
JOIN Arco a ON s.id_arco = a.id;
""",
# b)
"""
SELECT m.data, i.nome AS nome_inquinante, m.value AS valore_misurato, i.soglia_attenzione, i.soglia_allarme, s.nome AS nome_stazione
FROM Misurazione m
JOIN Inquinante i ON m.id_inquinante = i.id
JOIN StazioneDiMisurazione s ON m.id_stazione = s.id
WHERE DATE(m.data) = '2024-04-01'
ORDER BY i.nome, m.data, s.nome;
""",
# c)
"""
SELECT a.id AS id_arco, a.toponimo AS nome_arco, AVG(m.value)/a.capacita_max AS livello_di_servizio
FROM Misurazione m
JOIN StazioneDiMisurazione s ON m.id_stazione = s.id
JOIN Arco a ON s.id_arco = a.id
WHERE DATE(m.data) = '2024-04-01' AND s.tipo = 'traffico'
GROUP BY a.id;
""",
# d)
"""
SELECT s.id AS id_stazione, s.nome AS nome_stazione, COUNT(*) as numero_allarmi
FROM StazioneDiMisurazione s
JOIN Misurazione m ON s.id = m.id_stazione
JOIN Inquinante i ON i.id = m.id_inquinante
WHERE s.tipo = "aria" AND m.value > i.soglia_allarme
AND m.data BETWEEN '2024-04-01 00:00:00' AND '2024-04-03 23:59:59'
GROUP BY s.id;
""",
# e)
None, # non corretta automaticamente
]

DEBUG = False

################################
# FINE PARTE DI CONFIGURAZIONE #
################################


def log(x):
    if DEBUG:
        print(x)

import mariadb
from openpyxl import load_workbook

# connessione al db
connection = mariadb.connect(
    host=DB_HOST,
    user=DB_USER,
    database=DB_NAME,
    password=DB_PASS
)
cursor = connection.cursor()
print("Connessione al db riuscita")

# Confronta la risposta con quella attesa, restituisce:
# * 1 = risposta giusta
# * lettera = risposta errata (o parzialmente errata)
#     C => errore nella query corretta
#     A => errore nella query dello studente
#     L => lunghezza risultati query non corretta
#     E => risultati query non corretti
def compare_answers(correct, attempt):
    if correct is None:
        return "C"
    if attempt is None:
        return "A"

    def try_query(query, errcode):
        try:
            cursor.execute(query)
        except:
            log(f"ERROR RUNNING QUERY:\n{query}")
            return (None, errcode)
        try:
            res = cursor.fetchall()
        except:
            log(f"ERROR FETCHING RESULTS:\n{query}")
            return (None, f"{errcode}F")
        return (res, None)

    (cres, err) = try_query(correct, "C")
    if err:
        return err

    (res, err) = try_query(attempt, "A")
    if err:
        return err

    if len(cres) != len(res):
        return "L"

    for (cr, r) in zip(cres, res):
        if (cr != r):
            return "E"

    return 1

# dalla mail tira fuori cognome e iniziale del nome
# formati mail: "mrossi@qualcosa.it" o "rossi.m@qualcosa.it
def mail_to_name(mail):
    name = mail.split('@')[0]
    name_split = name.split(".")
    if (len(name_split) > 1):
        lastn = name_split[0]
        firstn = name_split[1]
    else:
        lastn = name[1:]
        firstn = name[0]
    return [ firstn, lastn ]


# Load the workbook
wb = load_workbook(FILE_NAME)

# Get a specific sheet by name
sheet = wb[SHEET_NAME]
if OUTPUT_SHEET_NAME in wb:
    out = wb[OUTPUT_SHEET_NAME]
else:
    out = wb.create_sheet(OUTPUT_SHEET_NAME)

i = 0
cell_value = sheet.cell(row=2, column=2).value
while cell_value is not None:
    i += 1
    cell_value = sheet.cell(row=2+i, column=2).value
student_count = i

for j in range(len(ANSWERS)):
    out.cell(row=1, column=3+j).value = f"#{j}"


for i in range(student_count):
    student_mail = sheet.cell(row=2+i, column=2).value
    [ firstn, lastn ] = mail_to_name(student_mail)
    out.cell(row=2+i, column=1).value = lastn
    out.cell(row=2+i, column=2).value = firstn
    print(f"Correggendo: {lastn} {firstn}")

    for j in range(len(ANSWERS)):
        student_answer = sheet.cell(row=2+i, column=3+j).value
        res = compare_answers(ANSWERS[j], student_answer)
        out.cell(row=2+i, column=3+j).value = res

wb.save(FILE_NAME)

print("Corretto tutto e aggiornato il file excel ;)")


